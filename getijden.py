import sys
from pathlib import Path
import re
import json
import datetime
import dateparser
from ics import Calendar, Event

from openpyxl import load_workbook

if len(sys.argv) < 2:
  print("ERROR: please provide one or more folders with tide XLSX files.")
  sys.exit(1)

folders = [ Path(folder) for folder in sys.argv[1:] ]

trg_folder = Path("ics")

def index(chars):
  return [ ord(char[0]) - ord("A") for char in chars ]

months = {
  "M1": {
    "first" : 4,
    "day"   : index([ "A", "H" ]),
    "high"  : index([ "C", "J" ]),
    "low"   : index([ "E", "L" ])
  },
  "AA1": {
    "first" : 4,
    "day"   : index([ "O", "V" ]),
    "high"  : index([ "Q", "X" ]),
    "low"   : index([ "S", "Z" ])
  }
}

TIME_FORMAT="%H:%M"

def add(cal, date, high, low):
  # construct summary
  summary = "🌊"

  if isinstance(high, datetime.time):
    high = datetime.datetime.combine(date.date(), high)
    summary += f" 🔺 {high.strftime(TIME_FORMAT)}"
  else:
    high = None

  if isinstance(low, datetime.time):
    low = datetime.datetime.combine(date.date(), low)
    summary += f" 🔻 {low.strftime(TIME_FORMAT)}"
  else:
    low = None

  # add event
  if high or low:
    event = Event(name=summary, begin=high if high else low)
    event.make_all_day()
    cal.events.add(event)

def determine_location(xlsx):
  # the filename change over time
  #   antwerpen2024-getijtabel-dmLAT.xlsx
  #   Antwerpen2025_dmLAT.xlsx
  #   ^^^^^^^^^
  location = re.split(r'[0-9]+', xlsx.stem)[0].lower()

  # location names also sometimes change
  try:
    location = {
      "wintam" : "wintamsluis"
    }[location]
  except KeyError:
    pass

  return location

def load(xlsx, cal):
  wb = load_workbook(filename=xlsx)

  # process all worksheets
  for ws in wb.worksheets:
    # extract each month according to configuration
    for month_cell, params in months.items():
      month_year = ws[month_cell].value

      # track previous day to fill in the blanks
      prev_day = [ None for _ in params["day"] ]
      
      # process all rows
      for row in ws.iter_rows(min_row=params["first"], values_only=True):

        # extract each set of day-based columns
        for index, cols in enumerate(zip(params["day"], params["high"], params["low"])):
          day, high, low = [ row[col] for col in cols ]

          # parse day, reusing if omitted
          try:
            day = int(day)
            prev_day[index] = day
          except (ValueError, TypeError):
            day = prev_day[index]

          # create date
          date = dateparser.parse(f"{day} {month_year}")

          # create event
          add(cal, date, high, low)
          
# load xlsx files into calendars

cals = {}
for folder in folders:
  print(f"🗂️ loading tide information from {folder}")
  for xlsx in folder.glob("*.xlsx"):
    location = determine_location(xlsx)
    print(f"👷‍♂️ processing {location}")
    try:
      cal = cals[location]
    except KeyError:
      cal = Calendar(creator="https://github.com/christophevg/getijden")
      cals[location] = cal
    load(xlsx, cal)

# create ICS files
trg_folder.mkdir(parents=True, exist_ok=True)
for name, cal in cals.items():
  trg_file = trg_folder / Path(name).with_suffix(".ics")
  print(f"📅 creating {trg_file}")
  with open( trg_file, "w") as f:
    f.write(cal.serialize())
